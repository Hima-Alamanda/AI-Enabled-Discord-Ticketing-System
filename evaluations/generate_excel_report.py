import pandas as pd
import os
import warnings
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

# --- CONFIGURATION ---
BASE_PATH = "/Users/himanth/PCB/main_ST_AI/Main/AI-Enabled-Ticketing-System-main 6/evaluations/results/"
INPUT_CSV = os.path.join(BASE_PATH, "prompt_comparison.csv")
OUTPUT_EXCEL = os.path.join(BASE_PATH, "EXECUTIVE_PERFORMANCE_REPORT_V7_final.xlsx")

# Suppress pandas warnings
warnings.filterwarnings('ignore')

def apply_header_style(cell):
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Side(style='thin', color="000000")
    cell.border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)

def section_header_style(cell):
    cell.font = Font(bold=True, size=12, color="1F4E78")
    cell.alignment = Alignment(horizontal="left", vertical="center")

def generate_professional_report():
    if not os.path.exists(INPUT_CSV):
        print(f"Error: {INPUT_CSV} not found.")
        return
    
    print("Processing Data for Multi-Sheet Official Report...")
    # Read the CSV - we use low_memory=False to handle any mixed types, and skip errors
    df_raw = pd.read_csv(INPUT_CSV, on_bad_lines='skip')
    
    # 1. CLEANUP: Remove any legacy footer lines if they exist in the CSV (from manual edits)
    # We identify rows that aren't real results (e.g., if Prompt_Strategy contains "Quality Scores")
    df_raw = df_raw[df_raw['Prompt_Strategy'].astype(str).str.contains('PROMPT', na=False)]
    
    df_raw.columns = [c.strip() for c in df_raw.columns]

    # Map current headers to standard internal names
    col_map = {}
    for col in df_raw.columns:
        if 'Correctness' in col: col_map['Correctness'] = col
        if 'Faithfulness' in col: col_map['Faithfulness'] = col
        if 'Actionability' in col: col_map['Actionability'] = col
        if 'Format' in col: col_map['Format'] = col
        if 'Ambiguity' in col: col_map['Ambiguity'] = col
        if 'Multimodal' in col: col_map['Multimodal'] = col
        if 'Escalation' in col: col_map['Escalation'] = col
        if 'Empathy' in col: col_map['Empathy'] = col
        if 'Latency' in col: col_map['Latency'] = col
        if 'BLEU' in col: col_map['BLEU'] = col
        if 'ROUGE' in col: col_map['ROUGE'] = col
        if 'BERTScore' in col: col_map['BERTScore'] = col
        if 'In_Tokens' in col: col_map['In_Tokens'] = col
        if 'Out_Tokens' in col: col_map['Out_Tokens'] = col
        if 'Total_Tokens' in col: col_map['Total_Tokens'] = col

    # Define the display name mapping with scales
    DISPLAY_NAME_MAP = {
        'Correctness': 'Correctness (1-5)',
        'Faithfulness': 'Faithfulness (1-5)',
        'Actionability': 'Actionability (1-5)',
        'Format': 'Format Adherence (1-5)',
        'Ambiguity': 'Ambiguity Handling (1-5)',
        'Multimodal': 'Multimodal (1-5)',
        'Escalation': 'Escalation Logic (1-5)',
        'Empathy': 'Empathy & Tone (1-5)',
        'BLEU': 'BLEU Score (0-1)',
        'ROUGE': 'ROUGE-L (0-1)',
        'BERTScore': 'BERTScore (0-1)',
        'In_Tokens': 'In_Tokens (1M-2M)',
        'Out_Tokens': 'Out_Tokens (30k-65k)',
        'Total_Tokens': 'Total_Tokens',
        'Latency': 'Latency (ms)'
    }

    # Convert numeric
    SCORE_COLS = [col_map.get(k) for k in ['Correctness', 'Faithfulness', 'Actionability', 'Format', 'Ambiguity', 'Multimodal', 'Escalation', 'Empathy'] if col_map.get(k)]
    ACCURACY_COLS = [col_map.get(k) for k in ['BLEU', 'ROUGE', 'BERTScore'] if col_map.get(k)]
    TOKEN_COLS = [col_map.get(k) for k in ['In_Tokens', 'Out_Tokens', 'Total_Tokens'] if col_map.get(k)]
    LATENCY_COL = col_map.get('Latency')

    all_numeric = SCORE_COLS + ACCURACY_COLS + TOKEN_COLS + ([LATENCY_COL] if LATENCY_COL else [])
    for c in all_numeric:
        df_raw[c] = pd.to_numeric(df_raw[c], errors='coerce').fillna(0)

    # Apply Display Name Mapping
    rename_dict = {col_map[k]: DISPLAY_NAME_MAP[k] for k in col_map if k in DISPLAY_NAME_MAP}
    df_raw.rename(columns=rename_dict, inplace=True)
    
    # Update score list names for downstream logic
    SCORE_COLS = [DISPLAY_NAME_MAP[k] for k in ['Correctness', 'Faithfulness', 'Actionability', 'Format', 'Ambiguity', 'Multimodal', 'Escalation', 'Empathy'] if k in DISPLAY_NAME_MAP]
    ACCURACY_COLS = [DISPLAY_NAME_MAP[k] for k in ['BLEU', 'ROUGE', 'BERTScore'] if k in DISPLAY_NAME_MAP]
    TOKEN_COLS = [DISPLAY_NAME_MAP[k] for k in ['In_Tokens', 'Out_Tokens', 'Total_Tokens'] if k in DISPLAY_NAME_MAP]
    LATENCY_COL_DISPLAY = DISPLAY_NAME_MAP.get('Latency')

    # --- SHEET 1: RAW FULL DATA ---
    raw_sheet_df = df_raw.copy()
    
    # --- SHEET 2: AVG QUALITY RESULTS ---
    quality_avg = df_raw.groupby(['Prompt_Strategy', 'Model']).agg({c: 'mean' for c in SCORE_COLS + ACCURACY_COLS + ([LATENCY_COL_DISPLAY] if LATENCY_COL_DISPLAY in df_raw.columns else [])}).reset_index()
    if LATENCY_COL_DISPLAY in quality_avg.columns:
        quality_avg.rename(columns={LATENCY_COL_DISPLAY: 'Avg Latency (s)'}, inplace=True)
        quality_avg['Avg Latency (s)'] = (quality_avg['Avg Latency (s)'] / 1000).round(2)

    # --- SHEET 3: TOKENS & COST ---
    tokens_df = df_raw[['Prompt_Strategy', 'Model', 'Case_ID'] + TOKEN_COLS].copy()

    # --- SHEET 4: EXECUTIVE GLOSSARY & DEFINITIONS ---
    summary_data = [
        ["I. EXECUTIVE SUMMARY & REPORT OVERVIEW", ""],
        ["Audit Context", "Comparing multi-model performance across disparate prompt engineering strategies."],
        ["", ""],
        ["II. PROMPT STRATEGY DEFINITIONS", ""],
        ["PROMPT A", "The current live system prompt. Acts as a Senior Support Engineer with a natural flow."],
        ["PROMPT B (Balanced)", "Optimized for a highly professional and empathetic helpdesk experience."],
        ["PROMPT C (High-Efficiency)", "Focused on pure technical resolution speed by removing all conversational filler."],
        ["PROMPT D (Cautious Diagnostic)", "Prioritizes thorough investigation and root-cause mapping before suggesting a fix."],
        ["", ""],
        ["III. QUALITY METRIC DEFINITIONS (Scale: 1-5)", ""],
        ["Score: 5", "Excellent: Perfectly correct, clear, and follows all PCB rules."],
        ["Score: 4", "Good: Minor wording or tone issues, technically sound."],
        ["Score: 3", "Acceptable: Mostly correct but missing some details or polish."],
        ["Score: 2", "Poor: Important mistakes, unclear steps, or ignored rules."],
        ["Score: 1", "Failure: Incorrect, misleading, or failed to answer."],
        ["", ""],
        ["1. Correctness", "Whether the bot gives the right answer based on past resolved tickets or known support information."],
        ["2. Faithfulness", "Whether the bot stays accurate and does not make up information."],
        ["3. Actionability", "Whether the bot gives clear steps that the user can actually follow."],
        ["4. Format Adherence", "Whether the bot follows the expected response structure or headings."],
        ["5. Ambiguity Handling", "Identifying vague queries and asking for details instead of guessing."],
        ["6. Multimodal", "Whether the bot can correctly use information from attached images or screenshots."],
        ["7. Escalation Logic", "Whether the bot correctly decides to solve the issue itself or hand it over to a human technician."],
        ["8. Empathy & Tone", "Whether the bot sounds professional, polite, and helpful."],
        ["", ""],
        ["IV. PERFORMANCE & COST SCALES", ""],
        ["BLEU / ROUGE (0 - 1)", "Text similarity scores. Higher means the response wording and structure are closer to the expected answer."],
        ["BERTScore (0 - 1)", "Semantic similarity score. Higher means the response matches the meaning and intent of the expected answer, even if different words are used."],
        ["Latency (ms / sec)", "Response time taken by the system to generate an answer. Lower is better. Raw values may be measured in milliseconds, while averages can be shown in seconds for readability."],
        ["Gemini 2.5 Pro", "Input: 1,000,000 Tokens | Max Output: ~65,535 Tokens"],
        ["Grok 4.20 Reasoning", "Input: 2,000,000 Tokens | Max Output: 30,000 Tokens"],
    ]
    summary_df = pd.DataFrame(summary_data)

    # --- SHEET 5: STRATEGIC MODEL RECOMMENDATIONS ---
    rec_data = [
        ["METRIC", "WINNING COMBINATION", "AVG SCORE", "STRATEGIC INSIGHT"],
    ]
    
    # Quality Columns
    metrics_to_eval = [
        ('Correctness (1-5)', 'Highest technical accuracy and logical soundness.'),
        ('Faithfulness (1-5)', 'Best at sticking to SOPs and avoiding hallucinations.'),
        ('Empathy & Tone (1-5)', 'Most natural, professional, and helpful tone.'),
        ('Actionability (1-5)', 'Clearest and most executable troubleshooting steps.'),
        ('BERTScore (0-1)', 'Best semantic alignment with the "Golden Answer" meaning.'),
    ]
    
    for metric_name, insight in metrics_to_eval:
        if metric_name in quality_avg.columns:
            top_row = quality_avg.loc[quality_avg[metric_name].idxmax()]
            rec_data.append([
                metric_name.split(' (')[0], 
                f"{top_row['Model']} [{top_row['Prompt_Strategy']}]",
                f"{top_row[metric_name]:.2f}",
                insight
            ])
            
    # Latency (Lower is better)
    if 'Avg Latency (s)' in quality_avg.columns:
        best_lat_row = quality_avg.loc[quality_avg['Avg Latency (s)'].idxmin()]
        rec_data.append([
            "Operational Speed",
            f"{best_lat_row['Model']} [{best_lat_row['Prompt_Strategy']}]",
            f"{best_lat_row['Avg Latency (s)']}s",
            "Fastest response time for real-time support."
        ])
    
    rec_df = pd.DataFrame(rec_data)

    # --- EXCEL GENERATION ---
    print(f"Creating Multi-Sheet Report: {OUTPUT_EXCEL}")
    with pd.ExcelWriter(OUTPUT_EXCEL, engine='openpyxl') as writer:
        raw_sheet_df.to_excel(writer, sheet_name="1_FULL_BENCHMARK_DATA", index=False)
        quality_avg.to_excel(writer, sheet_name="2_QUALITY_RESULTS_AVG", index=False)
        tokens_df.to_excel(writer, sheet_name="3_TOKEN_COST_ANALYSIS", index=False)
        summary_df.to_excel(writer, sheet_name="4_EXECUTIVE_GLOSSARY", index=False, header=False)
        rec_df.to_excel(writer, sheet_name="5_MODEL_RECOMMENDATIONS", index=False, header=False)

        workbook = writer.book

        # Styles
        score_rule = ColorScaleRule(start_type='num', start_value=0, start_color='F8696B', mid_type='num', mid_value=2.5, mid_color='FFEB84', end_type='num', end_value=5.0, end_color='63BE7B')
        acc_rule = ColorScaleRule(start_type='num', start_value=0, start_color='F8696B', mid_type='num', mid_value=0.5, mid_color='FFEB84', end_type='num', end_value=1.0, end_color='63BE7B')
        latency_rule = ColorScaleRule(start_type='num', start_value=10000, start_color='63BE7B', mid_type='num', mid_value=30000, mid_color='FFEB84', end_type='num', end_value=60000, end_color='F8696B')

        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            
            if sheet_name != "4_EXECUTIVE_GLOSSARY":
                for cell in ws[1]: apply_header_style(cell)
            
            # Format Sheet 4 & 5 Specifically
            if sheet_name in ["4_EXECUTIVE_GLOSSARY", "5_MODEL_RECOMMENDATIONS"]:
                for i, row in enumerate(ws.iter_rows(min_col=1, max_col=1)):
                    cell = row[0]
                    if cell.value and (any(cell.value.startswith(x) for x in ["I.", "II.", "III.", "IV."]) or cell.value == "METRIC"):
                        cell.font = Font(bold=True, size=13, color="1F4E78")
                        ws.row_dimensions[i+1].height = 20
                    elif cell.value:
                        cell.font = Font(bold=True, size=11)

            # Auto-Width
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for c_cell in col:
                    try:
                        if c_cell.value:
                            val_len = len(str(c_cell.value))
                            if val_len > max_length: max_length = val_len
                    except: pass
                ws.column_dimensions[column].width = min(max_length + 2, 70) if sheet_name == "4_EXECUTIVE_GLOSSARY" else min(max_length + 2, 40)

            # Conditional Formatting
            if sheet_name == "1_FULL_BENCHMARK_DATA":
                for idx, col_name in enumerate(raw_sheet_df.columns):
                    col_letter = get_column_letter(idx + 1)
                    range_str = f"{col_letter}2:{col_letter}{len(raw_sheet_df)+1}"
                    if col_name in SCORE_COLS: ws.conditional_formatting.add(range_str, score_rule)
                    elif col_name in ACCURACY_COLS: ws.conditional_formatting.add(range_str, acc_rule)
                    elif col_name == LATENCY_COL: ws.conditional_formatting.add(range_str, latency_rule)
            
            if sheet_name == "2_QUALITY_RESULTS_AVG":
                for idx, col_name in enumerate(quality_avg.columns):
                    col_letter = get_column_letter(idx + 1)
                    range_str = f"{col_letter}2:{col_letter}{len(quality_avg)+1}"
                    
                    # Apply Heatmaps and Number Formatting (2 Decimals)
                    if any(s in col_name for s in ['Correctness', 'Faithfulness', 'Actionability', 'Format', 'Ambiguity', 'Multimodal', 'Escalation', 'Empathy', 'BLEU', 'ROUGE', 'BERTScore']):
                        ws.conditional_formatting.add(range_str, acc_rule if any(x in col_name for x in ['Score', 'ROUGE', 'BERTScore']) else score_rule)
                        
                        # Apply 0.00 format to all rows in this column
                        for row_idx in range(2, len(quality_avg) + 2):
                            ws[f"{col_letter}{row_idx}"].number_format = '0.00'

    print(f"\nSUCCESS! Multi-Sheet Report Cleaned and Updated: {OUTPUT_EXCEL}")

if __name__ == "__main__":
    generate_professional_report()

